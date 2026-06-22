"""
Reportes/ReporteProduccion/exportadores.py

T038 — Generadores de archivos exportables: Excel, CSV, PDF.

Cada función recibe el dict `datos` que devuelve la vista
y retorna un HttpResponse listo para servir al cliente.
"""

import csv
import io
from datetime import date

from django.http import HttpResponse


def _fecha_corta(valor):
    if not valor:
        return ""
    texto = str(valor)
    parte = texto.split("T")[0].split(" ")[0]
    trozos = parte.split("-")
    if len(trozos) == 3:
        y, m, d = trozos
        if len(y) == 4:
            return f"{d}/{m}/{y}"
    return texto


def _detalle_por_tipo(datos: dict):
    tipo = datos.get("tipo_reporte", "resumen_general")
    detalle = datos.get("detalle", []) or []
    if tipo == "consultas":
        headers = ["Paciente", "CI", "Ficha", "Fecha consulta", "Médico", "Motivo", "CIE-10", "Diagnóstico", "Estado", "Nivel urgencia"]
        rows = [
            [
                r.get("paciente", ""),
                r.get("ci", ""),
                r.get("numero_ficha", ""),
                str(r.get("fecha_consulta", "")),
                r.get("medico", ""),
                r.get("motivo_consulta", ""),
                r.get("codigo_cie10", ""),
                r.get("diagnostico", ""),
                r.get("estado", ""),
                r.get("nivel_urgencia", ""),
            ]
            for r in detalle
        ]
        return "Consultas", headers, rows
    if tipo == "triajes":
        headers = ["Paciente", "Fecha", "Nivel urgencia", "FC", "FR", "Temperatura", "Saturación", "EVA", "Estado"]
        rows = [
            [
                r.get("paciente", ""),
                str(r.get("fecha", "")),
                r.get("nivel_urgencia", ""),
                r.get("fc", ""),
                r.get("fr", ""),
                r.get("temperatura", ""),
                r.get("saturacion", ""),
                r.get("eva", ""),
                r.get("estado", ""),
            ]
            for r in detalle
        ]
        return "Triajes", headers, rows
    if tipo in {"recetas", "recetas_emitidas", "recetas_dispensadas", "recetas_anuladas"}:
        headers = ["Número receta", "Fecha emisión", "Fecha dispensación", "Paciente", "Médico", "Dispensada por", "Estado", "Observaciones"]
        rows = [
            [
                r.get("numero_receta", ""),
                str(r.get("fecha_emision", "")),
                str(r.get("fecha_dispensacion", "")),
                r.get("paciente", ""),
                r.get("medico", ""),
                r.get("dispensada_por", ""),
                r.get("estado", ""),
                r.get("observaciones", "") or r.get("motivo_anulacion", ""),
            ]
            for r in detalle
        ]
        return "Recetas", headers, rows
    return "", [], []


def _detalle_consultas_general(datos: dict):
    detalle = datos.get("detalle_consultas", []) or []
    headers = ["Paciente", "CI", "Ficha", "Fecha", "Médico", "Motivo", "CIE-10", "Diagnóstico", "Estado"]
    rows = [
        [
            r.get("paciente", ""),
            r.get("ci", ""),
            r.get("numero_ficha", ""),
            _fecha_corta(r.get("fecha_consulta", "")),
            r.get("medico", ""),
            r.get("motivo_consulta", ""),
            r.get("codigo_cie10", ""),
            r.get("diagnostico", ""),
            r.get("estado", ""),
        ]
        for r in detalle
    ]
    return headers, rows


def _detalle_triajes_general(datos: dict):
    detalle = datos.get("detalle_triajes", []) or []
    headers = ["Paciente", "Fecha", "Nivel urgencia", "FC", "FR", "Temperatura", "Saturación", "EVA", "Estado"]
    rows = [
        [
            r.get("paciente", ""),
            _fecha_corta(r.get("fecha", "")),
            r.get("nivel_urgencia", ""),
            r.get("fc", ""),
            r.get("fr", ""),
            r.get("temperatura", ""),
            r.get("saturacion", ""),
            r.get("eva", ""),
            r.get("estado", ""),
        ]
        for r in detalle
    ]
    return headers, rows


def _detalle_recetas_general(datos: dict):
    detalle = datos.get("detalle_recetas", []) or []
    headers = ["Número receta", "Fecha emisión", "Paciente", "Médico", "Estado"]
    rows = [
        [
            r.get("numero_receta", ""),
            _fecha_corta(r.get("fecha_emision", "")),
            r.get("paciente", ""),
            r.get("medico", ""),
            r.get("estado", ""),
        ]
        for r in detalle
    ]
    return headers, rows


# ── Excel ─────────────────────────────────────────────────────────────────────

def exportar_excel(datos: dict) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    periodo   = datos["periodo"]
    nombre_archivo = f"reporte_produccion_{periodo['desde']}_{periodo['hasta']}.xlsx"

    # ── Estilos ───────────────────────────────────────────────────────────────
    estilo_titulo  = Font(bold=True, size=13, color="FFFFFF")
    estilo_header  = Font(bold=True, size=10, color="FFFFFF")
    fill_titulo    = PatternFill("solid", fgColor="1F4E79")
    fill_header    = PatternFill("solid", fgColor="2E75B6")
    fill_subtitulo = PatternFill("solid", fgColor="D6E4F0")
    centro         = Alignment(horizontal="center", vertical="center")

    def _hoja_tabla(nombre_hoja, titulo, encabezados, filas):
        ws = wb.create_sheet(title=nombre_hoja)

        # Título
        ws.merge_cells(f"A1:{get_column_letter(len(encabezados))}1")
        celda = ws["A1"]
        celda.value    = titulo
        celda.font     = estilo_titulo
        celda.fill     = fill_titulo
        celda.alignment = centro
        ws.row_dimensions[1].height = 22

        # Periodo
        ws.merge_cells(f"A2:{get_column_letter(len(encabezados))}2")
        celda2 = ws["A2"]
        celda2.value    = f"Periodo: {periodo['desde']} — {periodo['hasta']}"
        celda2.fill     = fill_subtitulo
        celda2.alignment = centro

        # Encabezados
        for col, texto in enumerate(encabezados, start=1):
            c = ws.cell(row=3, column=col, value=texto)
            c.font      = estilo_header
            c.fill      = fill_header
            c.alignment = centro

        # Datos
        for fila_idx, fila in enumerate(filas, start=4):
            for col_idx, valor in enumerate(fila, start=1):
                ws.cell(row=fila_idx, column=col_idx, value=valor)

        # Ajustar ancho de columnas
        for col in range(1, len(encabezados) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

        return ws

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────
    resumen = datos["resumen"]
    tipo = datos.get("tipo_reporte", "resumen_general")
    _hoja_tabla(
        "Resumen",
        "Resumen del Periodo",
        ["Indicador", "Valor"],
        [
            ["Total consultas",              resumen["total_consultas"]],
            ["Total triajes",                resumen["total_triajes"]],
            ["Recetas emitidas",             resumen["total_recetas_emitidas"]],
            ["Recetas dispensadas",          resumen["total_recetas_dispensadas"]],
            ["Recetas anuladas",             resumen["total_recetas_anuladas"]],
            ["Tasa de derivación (%)",       resumen["tasa_derivacion_pct"]],
        ],
    )
    if tipo == "resumen_general":
        _hoja_tabla(
            "Triajes por Nivel",
            "Distribucion de Triajes por Nivel de Urgencia",
            ["Nivel", "Total", "Porcentaje (%)"],
            [
                [r["nivel_urgencia"], r["total"], r["porcentaje"]]
                for r in datos["triajes_por_nivel"]
            ],
        )
        _hoja_tabla(
            "Produccion Medicos",
            "Produccion por Medico",
            ["Médico", "Consultas", "Recetas emitidas", "Derivaciones"],
            [
                [r["medico"], r["consultas"], r["recetas"], r["derivaciones"]]
                for r in datos["produccion_por_medico"]
            ],
        )
        _hoja_tabla(
            "Top Diagnosticos",
            "Top 10 Diagnosticos CIE-10",
            ["Código CIE-10", "Descripción", "Total consultas"],
            [
                [r["codigo"], r["descripcion"], r["total"]]
                for r in datos["top_diagnosticos"]
            ],
        )
        _hoja_tabla(
            "Consultas por Dia",
            "Volumen de Consultas por Dia",
            ["Fecha", "Total"],
            [[r["fecha"], r["total"]] for r in datos["consultas_por_dia"]],
        )
        headers_c, rows_c = _detalle_consultas_general(datos)
        _hoja_tabla(
            "Consultas",
            "Detalle de Consultas",
            headers_c,
            rows_c or [["No hay filas de detalle para exportar"] + [""] * (len(headers_c) - 1)],
        )
        headers_r, rows_r = _detalle_recetas_general(datos)
        _hoja_tabla(
            "Recetas",
            "Detalle de Recetas",
            headers_r,
            rows_r or [["No hay filas de detalle para exportar"] + [""] * (len(headers_r) - 1)],
        )
        headers_t, rows_t = _detalle_triajes_general(datos)
        _hoja_tabla(
            "Triajes",
            "Detalle de Triajes",
            headers_t,
            rows_t or [["No hay filas de detalle para exportar"] + [""] * (len(headers_t) - 1)],
        )
    elif tipo == "triajes":
        _hoja_tabla(
            "Triajes por Nivel",
            "Distribucion de Triajes por Nivel de Urgencia",
            ["Nivel", "Total", "Porcentaje (%)"],
            [
                [r["nivel_urgencia"], r["total"], r["porcentaje"]]
                for r in datos.get("triajes_por_nivel", [])
            ] or [["Sin datos", 0, 0]],
        )
        seccion, headers, rows = _detalle_por_tipo(datos)
        _hoja_tabla(
            seccion or "Detalle",
            f"Detalle de {seccion or 'Reporte'}",
            headers or ["Sin datos"],
            rows or [["No hay filas de detalle para exportar"]],
        )
    else:
        seccion, headers, rows = _detalle_por_tipo(datos)
        _hoja_tabla(
            seccion or "Detalle",
            f"Detalle de {seccion or 'Reporte'}",
            headers or ["Sin datos"],
            rows or [["No hay filas de detalle para exportar"]],
        )

    # ── Respuesta HTTP ────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


# ── CSV ───────────────────────────────────────────────────────────────────────

def exportar_csv(datos: dict) -> HttpResponse:
    periodo        = datos["periodo"]
    nombre_archivo = f"reporte_produccion_{periodo['desde']}_{periodo['hasta']}.csv"

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    response.write("﻿")  # BOM para Excel en Windows

    # Delimitador ';' para mejor apertura directa en Excel regional ES.
    writer = csv.writer(response, delimiter=";")
    resumen = datos["resumen"]
    filtros = datos.get("filtros_aplicados", {}) or {}

    # Encabezado
    writer.writerow(["REPORTE DE PRODUCCIÓN Y FLUJO DE ATENCIÓN — HISTOLINK"])
    writer.writerow([f"Periodo: {periodo['desde']} — {periodo['hasta']}"])
    if filtros:
        writer.writerow(["Filtros aplicados"])
        writer.writerow(["Filtro", "Valor"])
        for k, v in filtros.items():
            writer.writerow([k, v])
    writer.writerow([])

    tipo = datos.get("tipo_reporte", "resumen_general")
    writer.writerow(["Resumen"])
    writer.writerow(["Indicador", "Valor"])
    writer.writerow(["Total consultas",         resumen["total_consultas"]])
    writer.writerow(["Total triajes",            resumen["total_triajes"]])
    writer.writerow(["Recetas emitidas",         resumen["total_recetas_emitidas"]])
    writer.writerow(["Recetas dispensadas",      resumen["total_recetas_dispensadas"]])
    writer.writerow(["Recetas anuladas",         resumen["total_recetas_anuladas"]])
    writer.writerow(["Tasa de derivacion (%)",   resumen["tasa_derivacion_pct"]])
    writer.writerow([])
    if tipo == "resumen_general":
        writer.writerow(["Triajes por nivel"])
        writer.writerow(["Nivel", "Total", "Porcentaje (%)"])
        for r in datos["triajes_por_nivel"]:
            writer.writerow([r["nivel_urgencia"], r["total"], r["porcentaje"]])
        writer.writerow([])
        h_c, r_c = _detalle_consultas_general(datos)
        writer.writerow(["Detalle de consultas"])
        writer.writerow(h_c)
        for fila in r_c:
            writer.writerow(fila)
        writer.writerow([])
        h_r, r_r = _detalle_recetas_general(datos)
        writer.writerow(["Detalle de recetas"])
        writer.writerow(h_r)
        for fila in r_r:
            writer.writerow(fila)
        writer.writerow([])
        h_t, r_t = _detalle_triajes_general(datos)
        writer.writerow(["Detalle de triajes"])
        writer.writerow(h_t)
        for fila in r_t:
            writer.writerow(fila)
    elif tipo == "triajes":
        writer.writerow(["Triajes por nivel"])
        writer.writerow(["Nivel", "Total", "Porcentaje (%)"])
        for r in datos.get("triajes_por_nivel", []):
            writer.writerow([r.get("nivel_urgencia"), r.get("total"), r.get("porcentaje")])
        writer.writerow([])
        seccion, headers, rows = _detalle_por_tipo(datos)
        writer.writerow([f"Detalle {seccion}" if seccion else "Detalle"])
        writer.writerow(headers or ["Sin datos"])
        for r in rows:
            fila = list(r)
            for i, h in enumerate(headers or []):
                if "fecha" in str(h).lower():
                    fila[i] = _fecha_corta(fila[i])
            writer.writerow(fila)
    else:
        seccion, headers, rows = _detalle_por_tipo(datos)
        writer.writerow([])
        writer.writerow([f"Detalle {seccion}" if seccion else "Detalle"])
        if tipo == "consultas":
            headers = ["Paciente", "CI", "Ficha", "Fecha", "Médico", "Motivo", "CIE-10", "Diagnóstico", "Estado"]
            writer.writerow(headers)
            for r in rows:
                # filas consultas de _detalle_por_tipo:
                # Paciente, CI, Ficha, Fecha consulta, Médico, Motivo, CIE-10, Diagnóstico, Estado, Nivel urgencia
                fila = list(r[:9])
                fila[3] = _fecha_corta(fila[3])
                writer.writerow(fila)
        else:
            writer.writerow(headers or ["Sin datos"])
            for r in rows:
                fila = list(r)
                # Formateo básico de fechas para recetas/triajes si aplica
                for i, h in enumerate(headers or []):
                    if "fecha" in str(h).lower():
                        fila[i] = _fecha_corta(fila[i])
                writer.writerow(fila)

    return response


# ── PDF ───────────────────────────────────────────────────────────────────────

def exportar_pdf(datos: dict) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    periodo        = datos["periodo"]
    nombre_archivo = f"reporte_produccion_{periodo['desde']}_{periodo['hasta']}.pdf"
    resumen        = datos["resumen"]

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=1.5 * cm,
    )

    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "titulo", parent=estilos["Title"],
        fontSize=16, textColor=colors.HexColor("#1F4E79"), spaceAfter=4,
    )
    estilo_seccion = ParagraphStyle(
        "seccion", parent=estilos["Heading2"],
        fontSize=11, textColor=colors.HexColor("#2E75B6"), spaceBefore=12, spaceAfter=4,
    )

    COLOR_HEADER = colors.HexColor("#2E75B6")
    COLOR_FILA   = colors.HexColor("#EBF3FA")

    def _estilo_tabla(n_cols):
        return TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  COLOR_HEADER),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0),  8.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_FILA]),
            ("FONTSIZE",    (0, 1), (-1, -1), 7.5),
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("WORDWRAP",    (0, 0), (-1, -1), "CJK"),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])

    estilo_celda = ParagraphStyle("celda_pdf", parent=estilos["Normal"], fontSize=7.5, leading=8.8)
    estilo_celda_small = ParagraphStyle("celda_pdf_small", parent=estilos["Normal"], fontSize=7, leading=8)

    def _p(valor, small=False):
        return Paragraph(str(valor or ""), estilo_celda_small if small else estilo_celda)

    elementos = []

    # Título
    elementos.append(Paragraph(
        "Reporte de Producción y Flujo de Atención", estilo_titulo,
    ))
    elementos.append(Paragraph(
        f"Establecimiento: Histolink &nbsp;|&nbsp; Periodo: {periodo['desde']} — {periodo['hasta']}",
        estilos["Normal"],
    ))
    elementos.append(Spacer(1, 0.5 * cm))

    tipo = datos.get("tipo_reporte", "resumen_general")
    elementos.append(Paragraph("Resumen del Periodo", estilo_seccion))
    tabla_resumen = Table(
        [
            ["Indicador", "Valor"],
            ["Total consultas",          resumen["total_consultas"]],
            ["Total triajes",             resumen["total_triajes"]],
            ["Recetas emitidas",          resumen["total_recetas_emitidas"]],
            ["Recetas dispensadas",       resumen["total_recetas_dispensadas"]],
            ["Recetas anuladas",          resumen["total_recetas_anuladas"]],
            ["Tasa de derivación (%)",    resumen["tasa_derivacion_pct"]],
        ],
        colWidths=[12 * cm, 5 * cm],
    )
    tabla_resumen.setStyle(_estilo_tabla(2))
    elementos.append(tabla_resumen)
    elementos.append(Spacer(1, 0.4 * cm))
    if tipo == "resumen_general":
        elementos.append(Paragraph("Distribución de Triajes por Nivel", estilo_seccion))
        tabla_triajes_nivel = Table(
            [["Nivel", "Total", "Porcentaje (%)"]] +
            [[r["nivel_urgencia"], r["total"], r["porcentaje"]] for r in datos["triajes_por_nivel"]],
            colWidths=[5 * cm, 3.5 * cm, 4.5 * cm],
            repeatRows=1,
        )
        tabla_triajes_nivel.setStyle(_estilo_tabla(3))
        elementos.append(tabla_triajes_nivel)
        elementos.append(Spacer(1, 0.4 * cm))

        h_c, r_c = _detalle_consultas_general(datos)
        elementos.append(Paragraph("Detalle de Consultas", estilo_seccion))
        # Versión PDF reducida para legibilidad (sin Ficha ni Diagnóstico largo).
        h_c_pdf = ["Paciente", "CI", "Fecha", "Médico", "Motivo", "CIE-10", "Estado"]
        r_c_pdf = []
        for fila in r_c:
            r_c_pdf.append([
                _p(fila[0]),                    # Paciente
                _p(fila[1], small=True),        # CI
                _p(_fecha_corta(fila[3]), small=True),  # Fecha
                _p(fila[4]),                    # Médico
                _p(fila[5]),                    # Motivo
                _p(fila[6], small=True),        # CIE-10
                _p(fila[8], small=True),        # Estado
            ])
        body_c = [h_c_pdf] + (r_c_pdf if r_c_pdf else [[_p("No hay filas de detalle para exportar")] + [""] * (len(h_c_pdf) - 1)])
        tabla_c = Table(
            body_c,
            colWidths=[4.1 * cm, 1.9 * cm, 2.0 * cm, 3.3 * cm, 8.9 * cm, 1.7 * cm, 2.0 * cm],
            repeatRows=1,
        )
        tabla_c.setStyle(_estilo_tabla(len(h_c_pdf)))
        elementos.append(tabla_c)
        elementos.append(Spacer(1, 0.4 * cm))

        h_r, r_r = _detalle_recetas_general(datos)
        elementos.append(Paragraph("Detalle de Recetas Emitidas/Filtradas", estilo_seccion))
        body_r = [h_r] + ([[ _p(c) for c in row ] for row in r_r] if r_r else [[_p("No hay filas de detalle para exportar")] + [""] * (len(h_r) - 1)])
        tabla_r = Table(body_r, colWidths=[3.5 * cm, 2.4 * cm, 6.0 * cm, 5.2 * cm, 2.2 * cm], repeatRows=1)
        tabla_r.setStyle(_estilo_tabla(len(h_r)))
        elementos.append(tabla_r)
        elementos.append(Spacer(1, 0.4 * cm))

        h_t, r_t = _detalle_triajes_general(datos)
        elementos.append(Paragraph("Detalle de Triajes", estilo_seccion))
        body_t = [h_t] + ([[ _p(c) for c in row ] for row in r_t] if r_t else [[_p("No hay filas de detalle para exportar")] + [""] * (len(h_t) - 1)])
        tabla_t = Table(body_t, colWidths=[4.0 * cm, 2.2 * cm, 2.4 * cm, 1.5 * cm, 1.5 * cm, 2.2 * cm, 2.2 * cm, 1.2 * cm, 2.2 * cm], repeatRows=1)
        tabla_t.setStyle(_estilo_tabla(len(h_t)))
        elementos.append(tabla_t)
    elif tipo == "triajes":
        elementos.append(Paragraph("Distribución de Triajes por Nivel", estilo_seccion))
        tabla_triajes_nivel = Table(
            [["Nivel", "Total", "Porcentaje (%)"]] +
            [[r.get("nivel_urgencia"), r.get("total"), r.get("porcentaje")] for r in datos.get("triajes_por_nivel", [])],
            colWidths=[5 * cm, 3.5 * cm, 4.5 * cm],
            repeatRows=1,
        )
        tabla_triajes_nivel.setStyle(_estilo_tabla(3))
        elementos.append(tabla_triajes_nivel)
        elementos.append(Spacer(1, 0.4 * cm))
        seccion, headers, rows = _detalle_por_tipo(datos)
        elementos.append(Paragraph(f"Detalle de {seccion or 'Reporte'}", estilo_seccion))
        body = [headers] + ([[ _p(c) for c in row ] for row in rows] if rows else [[_p("No hay filas de detalle para exportar")]])
        cols = len(headers) if headers else 1
        col_widths = [24 * cm / cols for _ in range(cols)] if cols > 0 else [24 * cm]
        tabla = Table(body, colWidths=col_widths, repeatRows=1)
        tabla.setStyle(_estilo_tabla(cols))
        elementos.append(tabla)
    else:
        seccion, headers, rows = _detalle_por_tipo(datos)
        elementos.append(Paragraph(f"Detalle de {seccion or 'Reporte'}", estilo_seccion))
        if tipo == "consultas":
            # PDF legible para consultas: sin "Nivel urgencia", con wrapping y anchos fijos.
            headers_pdf = ["Paciente", "CI", "Fecha", "Médico", "Motivo", "CIE-10", "Estado"]
            rows_pdf = []
            for r in rows:
                fila = [
                    _p(r[0]),                    # Paciente
                    _p(r[1], small=True),        # CI
                    _p(_fecha_corta(r[3]), small=True),  # Fecha
                    _p(r[4]),                    # Medico
                    _p(r[5]),                    # Motivo
                    _p(r[6], small=True),        # CIE-10
                    _p(r[8], small=True),        # Estado
                ]
                rows_pdf.append(fila)
            body = [headers_pdf] + (rows_pdf if rows_pdf else [[_p("No hay filas de detalle para exportar")] + [""] * 6])
            col_widths = [4.1 * cm, 1.9 * cm, 2.0 * cm, 3.3 * cm, 8.9 * cm, 1.7 * cm, 2.0 * cm]
            tabla = Table(body, colWidths=col_widths, repeatRows=1)
            tabla.setStyle(_estilo_tabla(len(headers_pdf)))
            elementos.append(tabla)
        else:
            body = [headers] + (rows if rows else [["No hay filas de detalle para exportar"]])
            cols = len(headers) if headers else 1
            col_widths = [24 * cm / cols for _ in range(cols)] if cols > 0 else [24 * cm]
            tabla = Table(body, colWidths=col_widths, repeatRows=1)
            tabla.setStyle(_estilo_tabla(cols))
            elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


# ── SNIS — exportadores específicos ──────────────────────────────────────────

def exportar_snis_csv(datos: dict) -> HttpResponse:
    periodo        = datos["periodo"]
    nombre_archivo = f"reporte_snis_{periodo['fecha_desde']}_{periodo['fecha_hasta']}.csv"

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    response.write("﻿")  # BOM para Excel en Windows

    writer = csv.writer(response, delimiter=";")
    resumen = datos["resumen"]

    writer.writerow(["REPORTE SNIS — MORBILIDAD POR DIAGNÓSTICO CIE-10 — HISTOLINK"])
    writer.writerow([f"Período: {periodo['fecha_desde']} al {periodo['fecha_hasta']}"])
    writer.writerow([])
    writer.writerow(["RESUMEN"])
    writer.writerow(["Total de casos", resumen["total_casos"]])
    writer.writerow(["Diagnósticos distintos", resumen["total_diagnosticos_distintos"]])
    filtros = datos.get("filtros_aplicados", {})
    if filtros.get("codigo_cie10"):
        writer.writerow(["Filtro CIE-10", filtros["codigo_cie10"]])
    if filtros.get("sexo"):
        writer.writerow(["Filtro sexo", filtros["sexo"]])
    writer.writerow([])
    writer.writerow(["#", "Código CIE-10", "Descripción", "Total casos", "Masculino", "Femenino", "% del total"])
    for i, r in enumerate(datos["morbilidad"], start=1):
        writer.writerow([
            i,
            r["codigo"],
            r["descripcion"],
            r["total"],
            r["masculino"],
            r["femenino"],
            f"{r['porcentaje']}%",
        ])
    return response


def exportar_snis_excel(datos: dict) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    periodo        = datos["periodo"]
    nombre_archivo = f"reporte_snis_{periodo['fecha_desde']}_{periodo['fecha_hasta']}.xlsx"
    resumen        = datos["resumen"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SNIS Morbilidad"

    COLOR_TITULO  = "1F4E79"
    COLOR_HEADER  = "2E75B6"
    COLOR_SUB     = "D6E4F0"
    centro = Alignment(horizontal="center", vertical="center")

    # Título
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = "REPORTE SNIS — MORBILIDAD POR DIAGNÓSTICO CIE-10"
    t.font      = Font(bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=COLOR_TITULO)
    t.alignment = centro
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value     = f"Período: {periodo['fecha_desde']} — {periodo['fecha_hasta']}"
    s.fill      = PatternFill("solid", fgColor=COLOR_SUB)
    s.alignment = centro

    # Resumen
    ws["A3"] = "Total casos"
    ws["B3"] = resumen["total_casos"]
    ws["D3"] = "Diagnósticos distintos"
    ws["E3"] = resumen["total_diagnosticos_distintos"]
    for cell in [ws["A3"], ws["D3"]]:
        cell.font = Font(bold=True)

    # Cabecera tabla
    headers = ["#", "Código CIE-10", "Descripción", "Total", "Masculino", "Femenino", "% del total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        c.alignment = centro

    # Datos
    alt = PatternFill("solid", fgColor="EBF3FA")
    for i, r in enumerate(datos["morbilidad"], start=1):
        row = 5 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r["codigo"])
        ws.cell(row=row, column=3, value=r["descripcion"])
        ws.cell(row=row, column=4, value=r["total"])
        ws.cell(row=row, column=5, value=r["masculino"])
        ws.cell(row=row, column=6, value=r["femenino"])
        ws.cell(row=row, column=7, value=r["porcentaje"])
        if i % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = alt

    # Anchos de columna
    anchos = [5, 14, 40, 10, 12, 12, 12]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


def exportar_snis_pdf(datos: dict) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    periodo        = datos["periodo"]
    nombre_archivo = f"reporte_snis_{periodo['fecha_desde']}_{periodo['fecha_hasta']}.pdf"
    resumen        = datos["resumen"]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=1.5 * cm,
    )

    estilos       = getSampleStyleSheet()
    COLOR_HEADER  = colors.HexColor("#2E75B6")
    COLOR_TITULO  = colors.HexColor("#1F4E79")
    COLOR_FILA    = colors.HexColor("#EBF3FA")

    estilo_titulo = ParagraphStyle(
        "snis_titulo", parent=estilos["Title"],
        fontSize=15, textColor=COLOR_TITULO, spaceAfter=4,
    )
    estilo_seccion = ParagraphStyle(
        "snis_seccion", parent=estilos["Heading2"],
        fontSize=11, textColor=COLOR_HEADER, spaceBefore=10, spaceAfter=4,
    )
    estilo_celda = ParagraphStyle("snis_celda", parent=estilos["Normal"], fontSize=8, leading=9.5)

    def _p(valor):
        return Paragraph(str(valor or ""), estilo_celda)

    estilo_tabla = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  COLOR_HEADER),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  8.5),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, COLOR_FILA]),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])

    elementos = []

    elementos.append(Paragraph("Reporte SNIS — Morbilidad por Diagnóstico CIE-10", estilo_titulo))
    elementos.append(Paragraph(
        f"Establecimiento: Histolink &nbsp;|&nbsp; Período: {periodo['fecha_desde']} — {periodo['fecha_hasta']}",
        estilos["Normal"],
    ))
    filtros = datos.get("filtros_aplicados", {})
    partes_filtro = []
    if filtros.get("codigo_cie10"):
        partes_filtro.append(f"CIE-10: {filtros['codigo_cie10']}")
    if filtros.get("sexo"):
        partes_filtro.append(f"Sexo: {filtros['sexo']}")
    if partes_filtro:
        elementos.append(Paragraph("Filtros: " + " | ".join(partes_filtro), estilos["Normal"]))
    elementos.append(Spacer(1, 0.4 * cm))

    elementos.append(Paragraph("Resumen", estilo_seccion))
    tabla_resumen = Table(
        [
            ["Indicador", "Valor"],
            ["Total de casos registrados", resumen["total_casos"]],
            ["Diagnósticos distintos (CIE-10)", resumen["total_diagnosticos_distintos"]],
        ],
        colWidths=[10 * cm, 4 * cm],
    )
    tabla_resumen.setStyle(estilo_tabla)
    elementos.append(tabla_resumen)
    elementos.append(Spacer(1, 0.5 * cm))

    elementos.append(Paragraph("Morbilidad por Diagnóstico CIE-10", estilo_seccion))
    headers = ["#", "Código CIE-10", "Descripción del diagnóstico", "Total", "Masc.", "Fem.", "% del total"]
    filas = [headers]
    for i, r in enumerate(datos["morbilidad"], start=1):
        filas.append([
            i,
            _p(r["codigo"]),
            _p(r["descripcion"]),
            r["total"],
            r["masculino"],
            r["femenino"],
            f"{r['porcentaje']}%",
        ])
    if len(filas) == 1:
        filas.append(["—", "Sin registros para el período seleccionado", "", "", "", "", ""])

    tabla_morb = Table(
        filas,
        colWidths=[1.0 * cm, 3.0 * cm, 13.0 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm, 2.8 * cm],
        repeatRows=1,
    )
    tabla_morb.setStyle(estilo_tabla)
    elementos.append(tabla_morb)

    doc.build(elementos)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response
